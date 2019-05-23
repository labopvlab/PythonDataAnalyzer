#! python3

import os
from tkinter import filedialog
import xlsxwriter
from openpyxl import load_workbook
import six
import matplotlib.pyplot as plt

"""


"""


def QSSPCSummary():
    ready=0
    j=0
    while j<2:
        try: 
            file_path =filedialog.askopenfilenames(title="Please select the QSSPC files")
            if file_path!='':
                ready=1 
                directory = filedialog.askdirectory(title="Where saving?")
                    
                if not os.path.exists(directory):
                    os.makedirs(directory)
                    os.chdir(directory)
                else :
                    os.chdir(directory)
                break
            else:
                        print("Please select correct ellipso files")
                        j+=1
        except:
            print("no file selected")
            j+=1
    if ready:
        summaryA=[["Sample Name","Wafer Thickness (cm)","Resistivity (ohm-cm)","Optical constant","Specified MCD (cm^3)","Lifetime at Spec.MCD (us)","J0 (A/cm^2)","1 Sun Implied Voc (V)","Temp (DegC)"]]
        summaryB=[]
        totX=[]
        totY=[]
        legends=[]
        
        for i in range(len(file_path)):
            try:
                samplename=os.path.split(file_path[i])[1][:-5]
                index=samplename.index("_")
                batchname=samplename[:index]
                #measname=samplename[index+1:]
                
                wb = load_workbook(file_path[i], data_only=True)
                
                sheet=wb.get_sheet_by_name("User")
                
                WaferThickness=sheet['B6'].value
                resistivity=sheet['C6'].value
                Opticalconstant=sheet['E6'].value
                SpecifiedMCD=sheet['F6'].value
                LifetimeMCD=sheet['A9'].value
                J0=sheet['D9'].value
                OneSunImpliedVoc=sheet['K9'].value
                Temp=sheet['L9'].value
                summaryA.append([samplename,WaferThickness,resistivity,
                                Opticalconstant,SpecifiedMCD,LifetimeMCD,
                                J0,OneSunImpliedVoc,Temp])
                
                sheet=wb.get_sheet_by_name("RawData")
                x=[]
                y=[]
                Tau=["Tau (sec)",samplename]
                Minority=["Minority carrier density"," "]
                j=2
                while sheet['E'+str(j)].value!=None:
                    Tau.append(sheet['E'+str(j)].value)
                    y.append(float(sheet['E'+str(j)].value))
                    Minority.append(sheet['G'+str(j)].value)
                    x.append(sheet['G'+str(j)].value)
                    j+=1
                totX.append(x)
                totY.append(y)
                legends.append(samplename)
                plt.loglog(x,y,label=samplename)
                plt.xlabel('Minority carrier density')
                plt.ylabel('Lifetime')
                plt.legend()
                plt.savefig(samplename+'.png')
                plt.clf()
                summaryB.append(Minority)
                summaryB.append(Tau)
            except:
                print("some exception... pass")
        
        
        for i in range(len(legends)):
            plt.loglog(totX[i],totY[i],label=legends[i])
        plt.xlabel('Minority carrier density')
        plt.ylabel('Lifetime')
        plt.legend()
        plt.savefig(batchname+'.png')
        plt.clf()
        
        summaryB=list(map(list, six.moves.zip_longest(*summaryB, fillvalue=' ')))
        
        workbook = xlsxwriter.Workbook('QSSPC_summary.xlsx')
        
        worksheet = workbook.add_worksheet("SummaryA")
        for item in range(len(summaryA)):
            for item0 in range(len(summaryA[item])):
                worksheet.write(item,item0,summaryA[item][item0])
        
        worksheet = workbook.add_worksheet("SummaryB")
        for item in range(len(summaryB)):
            for item0 in range(len(summaryB[item])):
                worksheet.write(item,item0,summaryB[item][item0])
        
        workbook.close()

###############################################################################        
if __name__ == '__main__':
    QSSPCSummary()
